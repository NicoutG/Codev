from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from sqlalchemy.orm import Session
from typing import Dict, Any, List
from io import BytesIO
import json

class ExportService:
    def __init__(self, db: Session):
        self.db = db
    
    def export_indicator_results(self, results: List[Dict[str, Any]], 
                                 indicator_title: str, template_type: str = "default") -> BytesIO:
        """Export indicator results to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"
        
        # Header
        ws['A1'] = indicator_title
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Write data
        if results:
            # Headers
            headers = list(results[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data rows
            for row_idx, result in enumerate(results, start=3):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx).value = result.get(header)
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            # Get column letter safely
            try:
                column_letter = None
                for cell in column:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        break
                if not column_letter:
                    column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            except Exception:
                pass
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_formulaire(self, formulaire_data: Dict[str, Any], template_type: str = "CTI") -> BytesIO:
        """Export a complete formulaire with multiple indicators"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        from datetime import datetime
        ws_summary = wb.create_sheet("Résumé")
        ws_summary['A1'] = f"Formulaire: {formulaire_data.get('nom', 'N/A')}"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Demandeur: {formulaire_data.get('demandeur', 'N/A')}"
        ws_summary['A3'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Create sheet for each indicator
        indicators = formulaire_data.get('indicators', [])
        for idx, indicator_data in enumerate(indicators, start=1):
            ws = wb.create_sheet(f"Indicateur {idx}")
            self._write_indicator_to_sheet(ws, indicator_data, template_type)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _write_indicator_to_sheet(self, ws, indicator_data: Dict[str, Any], template_type: str):
        """Write indicator data to worksheet"""
        title = indicator_data.get('title', 'Indicateur')
        results = indicator_data.get('results', [])
        chart_type = indicator_data.get('chart_type', 'none')
        error = indicator_data.get('error')
        
        # Title
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        
        # If there's an error, show it
        if error:
            ws['A2'] = f"Erreur: {error}"
            ws['A2'].font = Font(size=12, color="FF0000", bold=True)
            return
        
        # Merge cells for title if we have results
        if results:
            num_cols = len(results[0].keys()) if results else 1
            from openpyxl.utils import get_column_letter
            end_col = get_column_letter(min(num_cols, 26))
            ws.merge_cells(f'A1:{end_col}1')
        
        if results:
            # Headers
            headers = list(results[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data
            for row_idx, result in enumerate(results, start=3):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx).value = result.get(header)
            
            # Auto-adjust widths
            for col_idx, column in enumerate(ws.columns, start=1):
                max_length = 0
                # Get column letter safely
                try:
                    # Try to get column letter from first non-merged cell
                    column_letter = None
                    for cell in column:
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                            break
                    # If still no column_letter, calculate from index
                    if not column_letter:
                        from openpyxl.utils import get_column_letter
                        column_letter = get_column_letter(col_idx)
                    
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                except Exception as e:
                    # Skip if there's an error with this column
                    pass
    
    def apply_cti_template(self, wb: Workbook):
        """Apply CTI template formatting"""
        # CTI specific formatting
        for ws in wb.worksheets:
            # Header style
            for row in ws.iter_rows(min_row=1, max_row=2):
                for cell in row:
                    if cell.value:
                        cell.font = Font(bold=True, size=12)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def apply_lyon1_template(self, wb: Workbook):
        """Apply Lyon 1 template formatting"""
        # Lyon 1 specific formatting
        for ws in wb.worksheets:
            # Different header style
            for row in ws.iter_rows(min_row=1, max_row=2):
                for cell in row:
                    if cell.value:
                        cell.font = Font(bold=True, size=11, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
