from fastapi import APIRouter, Depends, status, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime
import json
import io
import csv

from app.core.database import get_db
from app.api.deps import get_current_user, require_role
from app.models.user import User, UserRole
from app.schemas.report import (
    ReportCreate,
    ReportUpdate,
    ReportResponse,
    ReportGenerateResponse,
    IndicatorInReport
)
from app.services.report_service import ReportService
from app.services.indicator_execution_service import IndicatorExecutionService
from app.dao.report_dao import ReportDao

router = APIRouter()
service = ReportService()
execution_service = IndicatorExecutionService()
report_dao = ReportDao()


@router.get("/", response_model=List[ReportResponse])
def list_reports(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Liste tous les rapports. Accessible à tous les utilisateurs authentifiés."""
    reports = service.list_reports(db, skip=skip, limit=limit)
    return reports


@router.get("/{report_id}", response_model=ReportResponse)
def get_report(
    report_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Récupère un rapport avec ses indicateurs. Accessible à tous les utilisateurs authentifiés."""
    report = service.get_report(db, report_id)
    
    # Récupérer les indicateurs avec leur configuration
    indicators_with_config = report_dao.get_indicators_with_config(db, report_id)
    
    # Construire la réponse avec les indicateurs
    indicators_data = []
    for item in indicators_with_config:
        indicators_data.append(IndicatorInReport(
            id=item['indicator'].id,
            title=item['indicator'].title,
            description=item['indicator'].description,
            chart_type=item['chart_type'],
            chart_config=item['chart_config'],
            display_order=item['display_order']
        ))
    
    return {
        "id": report.id,
        "title": report.title,
        "description": report.description,
        "created_by": report.created_by,
        "created_at": report.created_at,
        "updated_at": report.updated_at,
        "indicators": indicators_data
    }


@router.post("/", response_model=ReportResponse, status_code=status.HTTP_201_CREATED)
def create_report(
    data: ReportCreate,
    current_user: User = Depends(require_role([UserRole.EDITEUR, UserRole.ADMIN])),
    db: Session = Depends(get_db),
):
    """Crée un nouveau rapport. Réservé aux éditeurs et admins."""
    return service.create_report(db, data, current_user)


@router.put("/{report_id}", response_model=ReportResponse)
def update_report(
    report_id: int,
    data: ReportUpdate,
    current_user: User = Depends(require_role([UserRole.EDITEUR, UserRole.ADMIN])),
    db: Session = Depends(get_db),
):
    """Modifie un rapport. Réservé aux éditeurs et admins."""
    return service.update_report(db, report_id, data)


@router.delete("/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_report(
    report_id: int,
    current_user: User = Depends(require_role([UserRole.EDITEUR, UserRole.ADMIN])),
    db: Session = Depends(get_db),
):
    """Supprime un rapport. Réservé aux éditeurs et admins."""
    service.delete_report(db, report_id)
    return None


@router.post("/{report_id}/generate", response_model=ReportGenerateResponse)
def generate_report(
    report_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Génère un rapport en exécutant tous ses indicateurs.
    Accessible à tous les utilisateurs authentifiés.
    """
    report = service.get_report(db, report_id)
    
    # Récupérer les indicateurs avec leur configuration
    indicators_with_config = report_dao.get_indicators_with_config(db, report_id)
    
    # Exécuter chaque indicateur
    results = []
    for item in indicators_with_config:
        indicator = item['indicator']
        try:
            # Créer une nouvelle session pour chaque indicateur pour éviter les transactions abortées
            from app.core.database import SessionLocal
            indicator_db = SessionLocal()
            try:
                execution_result = execution_service.execute_indicator(indicator_db, indicator.id)
                results.append({
                    "indicator_id": indicator.id,
                    "indicator_title": indicator.title,
                    "chart_type": item['chart_type'],
                    "execution_result": execution_result
                })
            finally:
                indicator_db.close()
        except Exception as e:
            # En cas d'erreur, on continue avec les autres indicateurs
            error_msg = str(e)
            # Extraire le détail de l'erreur HTTPException si disponible
            if hasattr(e, 'detail'):
                error_msg = e.detail
            results.append({
                "indicator_id": indicator.id,
                "indicator_title": indicator.title,
                "chart_type": item['chart_type'],
                "execution_result": {
                    "error": error_msg,
                    "rows": [],
                    "columns": [],
                    "row_count": 0
                }
            })
    
    return {
        "report_id": report.id,
        "report_title": report.title,
        "generated_at": datetime.now(),
        "results": results
    }


@router.get("/{report_id}/export")
def export_report(
    report_id: int,
    format: str = Query("json", regex="^(json|csv|excel)$", description="Format d'export: json, csv ou excel"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Exporte un rapport dans différents formats.
    Accessible à tous les utilisateurs authentifiés.
    """
    report = service.get_report(db, report_id)
    
    # Générer le rapport
    indicators_with_config = report_dao.get_indicators_with_config(db, report_id)
    
    results = []
    for item in indicators_with_config:
        indicator = item['indicator']
        try:
            execution_result = execution_service.execute_indicator(db, indicator.id)
            results.append({
                "indicator_id": indicator.id,
                "indicator_title": indicator.title,
                "chart_type": item['chart_type'],
                "execution_result": execution_result
            })
        except Exception as e:
            results.append({
                "indicator_id": indicator.id,
                "indicator_title": indicator.title,
                "chart_type": item['chart_type'],
                "execution_result": {
                    "error": str(e),
                    "rows": [],
                    "columns": [],
                    "row_count": 0
                }
            })
    
    if format == "json":
        # Export JSON
        export_data = {
            "report_id": report.id,
            "report_title": report.title,
            "generated_at": datetime.now().isoformat(),
            "results": results
        }
        return Response(
            content=json.dumps(export_data, indent=2, ensure_ascii=False),
            media_type="application/json",
            headers={
                "Content-Disposition": f'attachment; filename="report_{report.id}.json"'
            }
        )
    
    elif format == "csv":
        # Export CSV (tous les résultats concaténés)
        import csv
        import io
        
        sio = io.StringIO()
        writer = csv.writer(sio, delimiter=';', lineterminator='\n')
        
        # Écrire les résultats de chaque indicateur
        for result in results:
            if result['execution_result'].get('error'):
                continue
            
            exec_result = result['execution_result']
            writer.writerow([f"Indicateur: {result['indicator_title']}"])
            writer.writerow(exec_result['columns'])
            
            for row in exec_result['rows']:
                writer.writerow([row.get(col, '') for col in exec_result['columns']])
            
            writer.writerow([])  # Ligne vide entre indicateurs
        
        csv_text = sio.getvalue()
        csv_text = "\ufeff" + csv_text  # BOM UTF-8 pour Excel
        
        return Response(
            content=csv_text,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="report_{report.id}.csv"'
            }
        )
    
    elif format == "excel":
        # Export Excel (nécessite openpyxl)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            wb.remove(wb.active)  # Supprimer la feuille par défaut
            
            for result in results:
                if result['execution_result'].get('error'):
                    continue
                
                exec_result = result['execution_result']
                ws = wb.create_sheet(title=result['indicator_title'][:31])  # Limite Excel
                
                # Titre
                ws['A1'] = result['indicator_title']
                ws['A1'].font = Font(bold=True, size=14)
                
                # En-têtes
                for col_idx, col_name in enumerate(exec_result['columns'], start=1):
                    cell = ws.cell(row=3, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                
                # Données
                for row_idx, row in enumerate(exec_result['rows'], start=4):
                    for col_idx, col_name in enumerate(exec_result['columns'], start=1):
                        ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ''))
            
            # Sauvegarder dans un buffer
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return Response(
                content=buffer.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="report_{report.id}.xlsx"'
                }
            )
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Export Excel non disponible. Installez openpyxl: pip install openpyxl"
            )
    
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=f"Format d'export non supporté: {format}"
    )
